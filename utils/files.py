import csv
import openpyxl

def read_csv(path, encoding='utf-8'):
    encodings = [encoding, 'utf-8-sig', 'latin1', 'cp1252']
    last_exc = None
    for enc in encodings:
        try:
            with open(path, newline='', encoding=enc) as f:
                return list(csv.reader(f))
        except Exception as e:
            last_exc = e
    raise last_exc

def write_csv(path, rows, encoding='utf-8'):
    with open(path, 'w', newline='', encoding=encoding) as f:
        writer = csv.writer(f)
        writer.writerows(rows)

def read_txt(path, encoding='utf-8'):
    encodings = [encoding, 'utf-8-sig', 'latin1', 'cp1252']
    last_exc = None
    for enc in encodings:
        try:
            with open(path, encoding=enc) as f:
                return f.read()
        except Exception as e:
            last_exc = e
    raise last_exc

def write_txt(path, text, encoding='utf-8'):
    with open(path, 'w', encoding=encoding) as f:
        f.write(text)

def read_xlsx(path, sheet=0):
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[sheet]
    return [[cell.value for cell in row] for row in ws.iter_rows()]

def write_xlsx(path, data, sheet_name='Sheet1'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in data:
        ws.append(row)
    wb.save(path)